
import abbr as a
import urllib.parse, urllib.request, urllib.error
import xml.etree.ElementTree as ET
import json
from requests import Session
import pandas as pd

from datetime import datetime
import numpy as np
import openpyxl
import os



series = None
startDate = None
endDate = None 
aggregationTypes = None
formulas =None
frequency = None


class My_Manager:
    def __init__(self,_series,start_Date, end_Date, aggregation_Types,_formulas,_frequency):
        self.series = _series
        self.startDate = start_Date
        self.endDate =end_Date
        self.aggregationTypes =aggregation_Types
        self.formulas = _formulas
        self.frequency = _frequency
        
    def convertTime(self,timestamp):
        time = datetime.fromtimestamp(int(timestamp)).strftime('%d-%m-%y')
        return time


    def get_data(self):
        temp={}
        url = a.MAIN_PART
        path ='{}series={}&startDate={}&endDate={}&type=json&key={}&aggregationTypes={}&Formulas={}&Frequency={}'.format(url,self.series,self.startDate,self.endDate,a.API_KEY,self.aggregationTypes,self.formulas,self.frequency)
        
        #df_path = pd.read_json(path)
        #df = pd.DataFrame(data=df_path, columns=['Tarih','TP_DK_EUR_A_YTL'])
       
        url_to_open = urllib.request.urlopen(path).read()
        data = json.loads(url_to_open)
        
        new_series = self.series.replace('.','_') #URL'deki seri adı ile JSON sonuçundaki seri adı farklı geliyor.
        for i in data['items']:
            if i[new_series] == None:
                continue
            #temp['Time'] = self.convertTime(i['UNIXTIME']['$numberLong'])
            #temp['Time'] = (i['Tarih'])
            #temp['Value'] = i[new_series]
            temp[i['Tarih']] = i[new_series]
            #my_np =np.array([[temp['Time']], [temp['Value']]])
            
        #for x, y in temp.items():
        #    print(x, y)
  
        
        return temp
        
    def exportToExcel(self,data):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.delete_rows(idx=1)
        tempRow = 1
        for x in data.keys():
             tempCell = sheet.cell(row = tempRow, column = 1)
             tempCell.value = x
             tempRow += 1
             
        tempRow = 1     
        for y in data.values():
             tempCell = sheet.cell(row = tempRow, column = 2)
             tempCell.value = y
             tempRow += 1
          
        path = os.path.abspath(os.getcwd()) + "\\result.xlsx"
        wb.save(path)
       
        
  
        
        
        
            
        
        
        #if url_to_open.ok:
        #    pass
        #else:
        #    if url_to_open.status_code == 400:
        #        pass
        #    else:
        #        raise ConnectionError
        
        
        
        #url_to_open = urllib.request.urlopen(path).read()
        #tree = ET.fromstring(url_to_open)
        
        
        #url_to_open = urllib.request.urlopen(path).read()
        #tree = ET.parse(url_to_open)
        #root_node = tree.getroot()
        #lst = root_node.findall('TP_DK_USD_A_YTL')  
        